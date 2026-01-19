#!/usr/bin/env python3
"""
Etsy Shop Uploader Generator

Reads Amazon flatfile XLSM and generates Shop Uploader compatible XLSX.
Supports variations (Size & Color).
"""

import argparse
import logging
import sys
from pathlib import Path
from urllib.parse import quote

import openpyxl

# Force unbuffered output for real-time progress
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    stream=sys.stdout,
)
# Ensure stdout is unbuffered
sys.stdout.reconfigure(line_buffering=True) if hasattr(sys.stdout, 'reconfigure') else None

# Shop Uploader configuration (from your template's allowed_values)
SHOP_UPLOADER_CONFIG = {
    "category": "Signs (2844)",  # Signs category
    "shipping_profile": "Postage 2025 (208230423243)",
    "return_policy": "return=true|exchange=true|deadline=30 (1074420280634)",
    "processing_profile": 1402336022581,  # readiness_state_id from Etsy
}

# Size mappings from Amazon size codes to display and pricing
# Standard sizes: dracula, saville, dick, barzan, baby jesus
SIZE_CONFIG = {
    "XS": {
        "display": "95mm x 95mm",  # Same as S - smallest standard size
        "dimensions": (9.5, 9.5),  # cm
        "price": 10.99,
    },
    "S": {
        "display": "95mm x 95mm",
        "dimensions": (9.5, 9.5),  # cm
        "price": 10.99,
    },
    "M": {
        "display": "115mm x 95mm",
        "dimensions": (11.5, 9.5),
        "price": 12.99,
    },
    "L": {
        "display": "140mm x 90mm",
        "dimensions": (14.0, 9.0),
        "price": 14.99,
    },
    "XL": {
        "display": "194mm x 143mm",
        "dimensions": (19.4, 14.3),
        "price": 18.99,
    },
    "XXL": {
        "display": "290mm x 190mm",
        "dimensions": (29.0, 19.0),
        "price": 21.99,
    },
}

# Legacy mappings for backward compatibility
SIZE_DISPLAY = {k: v["display"] for k, v in SIZE_CONFIG.items()}
SIZE_DIMENSIONS = {k: v["dimensions"] for k, v in SIZE_CONFIG.items()}
SIZE_PRICES = {k: v["price"] for k, v in SIZE_CONFIG.items()}

# Shop Uploader column headers (from Template sheet row 1)
SHOP_UPLOADER_COLUMNS = [
    "listing_id",           # 1 - leave empty for new listings
    "parent_sku",           # 2 - groups variations
    "sku",                  # 3 - unique SKU
    "title",                # 4
    "description",          # 5
    "price",                # 6
    "quantity",             # 7
    "category",             # 8
    "_primary_color",       # 9
    "_secondary_color",     # 10
    "_occasion",            # 11
    "_holiday",             # 12
    "_deprecated_diameter", # 13
    "_deprecated_dimensions", # 14
    "_deprecated_fabric",   # 15
    "_deprecated_finish",   # 16
    "_deprecated_flavor",   # 17
    "_deprecated_height",   # 18
    "_deprecated_length",   # 19
    "_deprecated_material", # 20
    "_deprecated_pattern",  # 21
    "_deprecated_scent",    # 22
    "_deprecated_size",     # 23
    "_deprecated_style",    # 24
    "_deprecated_weight",   # 25
    "_deprecated_width",    # 26
    "_deprecated_device",   # 27
    "option1_name",         # 28
    "option1_value",        # 29
    "option2_name",         # 30
    "option2_value",        # 31
    "image_1",              # 32
    "image_2",              # 33
    "image_3",              # 34
    "image_4",              # 35
    "image_5",              # 36
    "image_6",              # 37
    "image_7",              # 38
    "image_8",              # 39
    "image_9",              # 40
    "image_10",             # 41
    "shipping_profile_id",  # 42
    "readiness_state_id",   # 43
    "return_policy_id",     # 44
    "length",               # 45
    "width",                # 46
    "height",               # 47
    "dimensions_unit",      # 48
    "weight",               # 49
    "weight_unit",          # 50
    "type",                 # 51
    "who_made",             # 52
    "is_made_to_order",     # 53
    "year_made",            # 54
    "is_vintage",           # 55
    "is_supply",            # 56
    "is_taxable",           # 57
    "auto_renew",           # 58
    "is_customizable",      # 59
    "is_personalizable",    # 60
    "personalization_is_required", # 61
    "personalization_instructions", # 62
    "personalization_char_count_max", # 63
    "style_1",              # 64
    "style_2",              # 65
    "tag_1",                # 66
    "tag_2",                # 67
    "tag_3",                # 68
    "tag_4",                # 69
    "tag_5",                # 70
    "tag_6",                # 71
    "tag_7",                # 72
    "tag_8",                # 73
    "tag_9",                # 74
    "tag_10",               # 75
    "tag_11",               # 76
    "tag_12",               # 77
    "tag_13",               # 78
    "action",               # 79
    "listing_state",        # 80
    "overwrite_images",     # 81
]


def read_amazon_flatfile(flatfile_path: Path) -> list[dict]:
    """
    Read Amazon flatfile XLSM and extract child product data.
    """
    logging.info("Opening flatfile: %s", flatfile_path.name)
    sys.stdout.flush()
    
    # Don't use read_only mode - it's slower for XLSM files with macros
    wb = openpyxl.load_workbook(flatfile_path, data_only=True)
    ws = wb["Template"]
    
    logging.info("Building column index...")
    sys.stdout.flush()
    
    # Build column index from row 3 (attribute names)
    col_index = {}
    for col in range(1, ws.max_column + 1):
        attr = ws.cell(3, col).value
        if attr:
            col_index[attr] = col
    
    products = []
    
    logging.info("Reading product rows...")
    sys.stdout.flush()
    
    # Read data rows (starting from row 4)
    # Stop when we hit empty rows
    empty_count = 0
    row = 4
    while empty_count < 5:  # Stop after 5 consecutive empty rows
        sku = ws.cell(row, col_index.get("item_sku", 2)).value
        
        if not sku:
            empty_count += 1
            row += 1
            continue
        
        empty_count = 0  # Reset counter
        parent_child = ws.cell(row, col_index.get("parent_child", 32)).value
        
        # Get parent SKU for variations
        parent_sku = ws.cell(row, col_index.get("parent_sku", 31)).value
        
        # Skip parent rows (we only need children)
        if parent_child == "Parent":
            row += 1
            continue
            
        product = {
            "sku": sku,
            "parent_sku": parent_sku or sku,  # Use own SKU if no parent
            "title": ws.cell(row, col_index.get("item_name", 10)).value,
            "description": ws.cell(row, col_index.get("product_description", 7)).value,
            "color": ws.cell(row, col_index.get("color_name", 45)).value,
            "size": ws.cell(row, col_index.get("size_name", 46)).value,
            "main_image": ws.cell(row, col_index.get("main_image_url", 13)).value,
            "image_2": ws.cell(row, col_index.get("other_image_url1", 14)).value,
            "image_3": ws.cell(row, col_index.get("other_image_url2", 15)).value,
            "image_4": ws.cell(row, col_index.get("other_image_url3", 16)).value,
            "image_5": ws.cell(row, col_index.get("other_image_url4", 17)).value,
            "keywords": ws.cell(row, col_index.get("generic_keywords", 44)).value,
        }
        
        products.append(product)
        row += 1
        
        if len(products) % 10 == 0:
            logging.info("Progress: Read %d products...", len(products))
            sys.stdout.flush()
    
    wb.close()
    logging.info("Read %d child products from %s", len(products), flatfile_path.name)
    sys.stdout.flush()
    return products


def convert_png_url_to_jpeg(url: str) -> str:
    """
    Convert a PNG URL to JPEG URL.
    Assumes JPEG version was uploaded alongside PNG during Amazon pipeline.
    """
    if url and url.endswith(".png"):
        return url[:-4] + ".jpg"
    return url


def encode_image_url(url: str) -> str:
    """
    URL-encode image URLs to handle spaces and special characters.
    Also converts PNG URLs to JPEG for Etsy compatibility.
    Only encodes the filename part, preserving the base URL.
    """
    if not url:
        return ""
    
    # Convert PNG to JPEG URL (JPEG was uploaded during Amazon pipeline)
    url = convert_png_url_to_jpeg(url)
    
    # Split URL into base and filename
    if "/" in url:
        base = url.rsplit("/", 1)[0]
        filename = url.rsplit("/", 1)[1]
        # Encode the filename (spaces become %20)
        encoded_filename = quote(filename, safe="")
        return f"{base}/{encoded_filename}"
    
    return url


def convert_keywords_to_tags(keywords: str) -> list[str]:
    """
    Convert Amazon keywords to Etsy tags (max 13 tags, each max 20 chars).
    """
    if not keywords:
        return []
    
    words = keywords.replace(",", " ").split()
    
    tags = []
    current_tag = ""
    
    for word in words:
        word = word.strip()
        if not word:
            continue
        
        if len(current_tag) == 0:
            current_tag = word
        elif len(current_tag) + 1 + len(word) <= 20:
            current_tag += " " + word
        else:
            if current_tag:
                tags.append(current_tag[:20])
            current_tag = word
        
        if len(tags) >= 13:
            break
    
    if current_tag and len(tags) < 13:
        tags.append(current_tag[:20])
    
    return tags[:13]


def generate_shop_uploader_file(
    products: list[dict],
    output_path: Path,
    template_path: Path,
    price: float = 9.99,
    quantity: int = 999,
    action: str = "create",
    listing_state: str = "draft",
) -> None:
    """
    Generate Shop Uploader XLSX from product data.
    """
    logging.info("Generating Shop Uploader file with %d products...", len(products))
    
    # Create fresh workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Write headers (row 1)
    for col_idx, col_name in enumerate(SHOP_UPLOADER_COLUMNS, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Write product data starting at row 2
    row_num = 2
    total = len(products)
    
    for idx, product in enumerate(products, 1):
        # Log progress
        if idx % 5 == 0 or idx == total:
            logging.info("Progress: %d/%d products", idx, total)
        
        size_code = product["size"]
        size_display = SIZE_DISPLAY.get(size_code, size_code)
        dimensions = SIZE_DIMENSIONS.get(size_code, (11.5, 9.5))
        # Use size-based pricing, fall back to provided price
        item_price = SIZE_PRICES.get(size_code, price)
        tags = convert_keywords_to_tags(product["keywords"])
        
        # Build row data
        row_data = {
            "listing_id": "",  # Empty for new listings
            "parent_sku": product["parent_sku"],
            "sku": product["sku"],
            "title": product["title"][:140] if product["title"] else "",
            "description": product["description"] or "",
            "price": item_price,
            "quantity": quantity,
            "category": SHOP_UPLOADER_CONFIG["category"],
            "_primary_color": product["color"],
            "option1_name": "Size",
            "option1_value": size_display,
            "option2_name": "_primary_color",
            "option2_value": product["color"],
            "image_1": encode_image_url(product["main_image"]),
            "image_2": encode_image_url(product["image_2"]),
            "image_3": encode_image_url(product["image_3"]),
            "image_4": encode_image_url(product["image_4"]),
            "image_5": encode_image_url(product["image_5"]),
            "shipping_profile_id": SHOP_UPLOADER_CONFIG["shipping_profile"],
            "readiness_state_id": SHOP_UPLOADER_CONFIG["processing_profile"],
            "return_policy_id": SHOP_UPLOADER_CONFIG["return_policy"],
            "length": dimensions[0],
            "width": dimensions[1],
            "height": 0.1,
            "dimensions_unit": "cm",
            "weight": 50,
            "weight_unit": "g",
            "type": "physical",
            "who_made": "i_did",
            "is_made_to_order": "TRUE",
            "year_made": "",
            "is_vintage": "FALSE",
            "is_supply": "FALSE",
            "is_taxable": "TRUE",
            "auto_renew": "TRUE",
            "is_customizable": "FALSE",
            "is_personalizable": "FALSE",
            "personalization_is_required": "FALSE",
            "style_1": "Modern",
            "action": action,
            "listing_state": listing_state,
            "overwrite_images": "TRUE",
        }
        
        # Add tags
        for i, tag in enumerate(tags, 1):
            row_data[f"tag_{i}"] = tag
        
        # Write to row
        for col_idx, col_name in enumerate(SHOP_UPLOADER_COLUMNS, 1):
            value = row_data.get(col_name, "")
            ws.cell(row=row_num, column=col_idx, value=value)
        
        row_num += 1
    
    # Save output
    wb.save(output_path)
    wb.close()
    
    logging.info("Generated %d listings to %s", len(products), output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate Shop Uploader file from Amazon flatfile")
    parser.add_argument("--input", type=Path, required=True, help="Input Amazon flatfile XLSM")
    parser.add_argument("--output", type=Path, default=None, help="Output Shop Uploader XLSX")
    parser.add_argument("--template", type=Path, 
                       default=Path(__file__).parent / "003 FLATFILES" / "Shop_Uploader_Template.xlsx",
                       help="Shop Uploader template XLSX (for reference)")
    parser.add_argument("--price", type=float, default=9.99, help="Listing price in GBP")
    parser.add_argument("--quantity", type=int, default=999, help="Stock quantity")
    parser.add_argument("--action", type=str, default="create", choices=["create", "update"], help="Action type")
    parser.add_argument("--state", type=str, default="draft", choices=["draft", "published"], help="Listing state")
    args = parser.parse_args()
    
    if not args.input.exists():
        logging.error("Input file not found: %s", args.input)
        return 1
    
    # Default output name
    if args.output is None:
        args.output = args.input.parent / f"{args.input.stem}_shop_uploader.xlsx"
    
    # Read Amazon flatfile
    products = read_amazon_flatfile(args.input)
    
    if not products:
        logging.error("No products found in flatfile")
        return 1
    
    # Generate Shop Uploader file
    generate_shop_uploader_file(
        products,
        args.output,
        args.template,
        price=args.price,
        quantity=args.quantity,
        action=args.action,
        listing_state=args.state,
    )
    
    logging.info("Done!")
    return 0


if __name__ == "__main__":
    exit(main())
