#!/usr/bin/env python3
"""
Convert transparent PNG images to JPEG with white background.

Downloads images from R2, converts them, and re-uploads with .jpg extension.
Updates the flatfile with new URLs.
"""

import argparse
import logging
import os
import re
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from urllib.parse import unquote

import boto3
import requests
from PIL import Image

# Number of parallel workers for image processing
MAX_WORKERS = 8

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)

# R2 Configuration (from environment variables set by config.bat)
R2_CONFIG = {
    "account_id": os.environ.get("R2_ACCOUNT_ID"),
    "access_key_id": os.environ.get("R2_ACCESS_KEY_ID"),
    "secret_access_key": os.environ.get("R2_SECRET_ACCESS_KEY"),
    "bucket_name": os.environ.get("R2_BUCKET_NAME"),
    "public_url": os.environ.get("R2_PUBLIC_URL"),
}


def get_s3_client():
    """Create S3 client for Cloudflare R2."""
    return boto3.client(
        "s3",
        endpoint_url=f"https://{R2_CONFIG['account_id']}.r2.cloudflarestorage.com",
        aws_access_key_id=R2_CONFIG["access_key_id"],
        aws_secret_access_key=R2_CONFIG["secret_access_key"],
        region_name="auto",  # R2 requires 'auto' as region
    )


def download_image(url: str, temp_dir: Path) -> Path:
    """Download image from URL to temp directory."""
    # Extract filename from URL (decode %20 etc)
    filename = unquote(url.split("/")[-1])
    local_path = temp_dir / filename
    
    response = requests.get(url, timeout=30)
    response.raise_for_status()
    
    with open(local_path, "wb") as f:
        f.write(response.content)
    
    return local_path


def convert_png_to_jpeg(png_path: Path, output_dir: Path, background_color=(255, 255, 255)) -> Path:
    """
    Convert PNG with transparency to JPEG with solid background.
    
    Args:
        png_path: Path to input PNG file
        output_dir: Directory for output JPEG
        background_color: RGB tuple for background (default white)
    
    Returns:
        Path to output JPEG file
    """
    # Open the PNG
    img = Image.open(png_path)
    
    # Strip ICC profile to avoid compatibility issues with Etsy
    img.info.pop('icc_profile', None)
    
    # Create output filename (change extension to .jpg)
    output_filename = png_path.stem + ".jpg"
    output_path = output_dir / output_filename
    
    # If image has alpha channel, composite onto background
    if img.mode in ("RGBA", "LA", "P"):
        # Create white background
        background = Image.new("RGB", img.size, background_color)
        
        # Convert to RGBA if needed
        if img.mode == "P":
            img = img.convert("RGBA")
        
        # Paste image onto background using alpha as mask
        if img.mode == "RGBA":
            background.paste(img, mask=img.split()[3])  # Use alpha channel as mask
        else:
            background.paste(img)
        
        img = background
    elif img.mode != "RGB":
        img = img.convert("RGB")
    
    # Save with optimized JPEG for Etsy compatibility
    img.save(output_path, "JPEG", quality=85, optimize=True)
    
    return output_path


def upload_to_r2(local_path: Path, s3_client) -> str:
    """Upload file to R2 and return public URL."""
    key = local_path.name
    
    s3_client.upload_file(
        str(local_path),
        R2_CONFIG["bucket_name"],
        key,
        ExtraArgs={"ContentType": "image/jpeg"},
    )
    
    public_url = f"{R2_CONFIG['public_url']}/{key}"
    logging.info("Uploaded %s -> %s", local_path.name, public_url)
    return public_url


def process_single_image(url: str, temp_path: Path, s3_client) -> tuple:
    """
    Process a single image: download, convert, upload.
    
    Returns:
        Tuple of (original_url, new_url) or (original_url, None) on failure
    """
    try:
        filename = url.split("/")[-1]
        logging.info("Processing %s", filename)
        
        # Download
        local_png = download_image(url, temp_path)
        
        # Convert
        local_jpg = convert_png_to_jpeg(local_png, temp_path)
        
        # Upload
        new_url = upload_to_r2(local_jpg, s3_client)
        
        # Clean up temp files
        local_png.unlink(missing_ok=True)
        local_jpg.unlink(missing_ok=True)
        
        return (url, new_url)
        
    except Exception as e:
        logging.error("Failed to process %s: %s", url, e)
        return (url, None)


def process_images_from_flatfile(flatfile_path: Path, output_flatfile: Path = None) -> dict:
    """
    Process all images from an Amazon flatfile using parallel processing.
    
    Returns:
        Dictionary mapping old PNG URLs to new JPEG URLs
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(flatfile_path, data_only=True)
    ws = wb["Template"]
    
    # Build column index
    col_index = {}
    for col in range(1, ws.max_column + 1):
        attr = ws.cell(3, col).value
        if attr:
            col_index[attr] = col
    
    # Collect all unique image URLs
    image_cols = ["main_image_url", "other_image_url1", "other_image_url2", 
                  "other_image_url3", "other_image_url4", "other_image_url5",
                  "other_image_url6", "other_image_url7", "other_image_url8"]
    
    all_urls = set()
    for row in range(4, ws.max_row + 1):
        for col_name in image_cols:
            if col_name in col_index:
                url = ws.cell(row, col_index[col_name]).value
                if url and url.endswith(".png"):
                    all_urls.add(url)
    
    wb.close()
    
    total_images = len(all_urls)
    logging.info("Found %d unique PNG images to convert (using %d parallel workers)", total_images, MAX_WORKERS)
    
    # Process images in parallel
    url_mapping = {}
    s3_client = get_s3_client()
    completed = 0
    
    with tempfile.TemporaryDirectory() as temp_dir:
        temp_path = Path(temp_dir)
        
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            # Submit all tasks
            futures = {
                executor.submit(process_single_image, url, temp_path, s3_client): url
                for url in sorted(all_urls)
            }
            
            # Collect results as they complete
            for future in as_completed(futures):
                original_url, new_url = future.result()
                if new_url:
                    url_mapping[original_url] = new_url
                completed += 1
                if completed % 10 == 0 or completed == total_images:
                    logging.info("Progress: %d/%d images processed", completed, total_images)
    
    return url_mapping


def update_flatfile_urls(flatfile_path: Path, url_mapping: dict, output_path: Path):
    """Update flatfile with new JPEG URLs."""
    import openpyxl
    
    wb = openpyxl.load_workbook(flatfile_path)
    ws = wb["Template"]
    
    # Build column index
    col_index = {}
    for col in range(1, ws.max_column + 1):
        attr = ws.cell(3, col).value
        if attr:
            col_index[attr] = col
    
    image_cols = ["main_image_url", "other_image_url1", "other_image_url2", 
                  "other_image_url3", "other_image_url4", "other_image_url5",
                  "other_image_url6", "other_image_url7", "other_image_url8"]
    
    updates = 0
    for row in range(4, ws.max_row + 1):
        for col_name in image_cols:
            if col_name in col_index:
                cell = ws.cell(row, col_index[col_name])
                if cell.value in url_mapping:
                    cell.value = url_mapping[cell.value]
                    updates += 1
    
    wb.save(output_path)
    wb.close()
    
    logging.info("Updated %d URLs in %s", updates, output_path)


def main():
    parser = argparse.ArgumentParser(description="Convert PNG images to JPEG with white background")
    parser.add_argument("--flatfile", type=Path, required=True, help="Amazon flatfile XLSM")
    parser.add_argument("--output", type=Path, default=None, help="Output flatfile with updated URLs")
    parser.add_argument("--url", type=str, default=None, help="Single URL to convert (for testing)")
    args = parser.parse_args()
    
    # Validate R2 config
    if not all(R2_CONFIG.values()):
        logging.error("R2 configuration incomplete. Run config.bat first.")
        logging.error("Missing: %s", [k for k, v in R2_CONFIG.items() if not v])
        return 1
    
    if args.url:
        # Single URL mode for testing
        s3_client = get_s3_client()
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            local_png = download_image(args.url, temp_path)
            local_jpg = convert_png_to_jpeg(local_png, temp_path)
            new_url = upload_to_r2(local_jpg, s3_client)
            print(f"Converted: {args.url}")
            print(f"New URL: {new_url}")
        return 0
    
    if not args.flatfile.exists():
        logging.error("Flatfile not found: %s", args.flatfile)
        return 1
    
    if args.output is None:
        args.output = args.flatfile.parent / f"{args.flatfile.stem}_jpeg{args.flatfile.suffix}"
    
    # Process all images
    url_mapping = process_images_from_flatfile(args.flatfile)
    
    if url_mapping:
        # Update flatfile with new URLs
        update_flatfile_urls(args.flatfile, url_mapping, args.output)
        logging.info("Done! Updated flatfile saved to %s", args.output)
    else:
        logging.warning("No images were converted")
    
    return 0


if __name__ == "__main__":
    exit(main())
