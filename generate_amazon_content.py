#!/usr/bin/env python3
"""
Amazon UK Flatfile Content Generator

Generates SEO-optimized Amazon product listings using Claude API,
uploads images to Cloudflare R2, and exports to Amazon flatfile format.
"""

import argparse
import csv
import json
import logging
import os
import re
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

# Number of parallel workers for image uploads
MAX_UPLOAD_WORKERS = 8

import anthropic
import boto3
from botocore.config import Config
import openpyxl
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)

# Product size dimensions in cm (length x width) - from PRIVATE SIGNAGE FLATFILE REV1
SIZE_DIMENSIONS_CM = {
    "dracula": (9.5, 9.5),      # XS - square
    "saville": (11.0, 9.5),     # S - landscape
    "dick": (14.0, 9.0),        # M - portrait
    "barzan": (19.0, 14.0),     # L - landscape
    "baby_jesus": (29.0, 19.0), # XL - portrait
    "a5": (21.0, 14.8),
    "a4": (29.7, 21.0),
    "a3": (42.0, 29.7),
}

# Color display names
COLOR_DISPLAY_NAMES = {
    "silver": "Silver",
    "white": "White",
    "gold": "Gold",
}

# Size display names (for flatfile output - customer-facing, use dimensions not internal names)
SIZE_DISPLAY_NAMES = {
    "dracula": "9.5 x 9.5 cm",
    "saville": "11 x 9.5 cm",
    "dick": "14 x 9 cm",
    "barzan": "19 x 14 cm",
    "baby_jesus": "29 x 19 cm",
    "a5": "A5 (21 x 14.8 cm)",
    "a4": "A4 (29.7 x 21 cm)",
    "a3": "A3 (42 x 29.7 cm)",
}

# Size display names with dimensions (for content generation prompts - internal use)
SIZE_DISPLAY_NAMES_FULL = {
    "dracula": "9.5 x 9.5 cm",
    "saville": "11 x 9.5 cm",
    "dick": "14 x 9 cm",
    "barzan": "19 x 14 cm",
    "baby_jesus": "29 x 19 cm",
    "a5": "A5 (21 x 14.8 cm)",
    "a4": "A4 (29.7 x 21 cm)",
    "a3": "A3 (42 x 29.7 cm)",
}

# Amazon size_map values (required for Size & Colour variation theme)
# Maps internal size names to Amazon's valid size codes: XXS, XS, S, M, L, XL, XXL
# From PRIVATE SIGNAGE FLATFILE REV1
SIZE_MAP_VALUES = {
    "dracula": "XS",     # 9.5 x 9.5 cm - Extra Small
    "saville": "S",      # 11 x 9.5 cm - Small
    "dick": "M",         # 14 x 9 cm - Medium
    "barzan": "L",       # 19 x 14 cm - Large
    "baby_jesus": "XL", # 29 x 19 cm - Extra Large
    "a5": "M",           # 21 x 14.8 cm
    "a4": "L",           # 29.7 x 21 cm
    "a3": "XL",          # 42 x 29.7 cm
}

# Pricing by size (from PRIVATE SIGNAGE FLATFILE REV1)
SIZE_PRICING = {
    "dracula": 10.99,    # XS
    "saville": 11.99,    # S
    "dick": 12.99,       # M
    "barzan": 15.99,     # L
    "baby_jesus": 17.99, # XL
}


@dataclass
class AmazonContent:
    """Generated Amazon listing content."""
    title: str
    description: str
    bullet_points: list[str]
    search_terms: str
    

# Material descriptions for content generation
MATERIALS = {
    "1mm_aluminium": "1mm brushed aluminium with UV-printed design",
    "3mm_aluminium_composite": "3mm aluminium composite with UV-printed design",
    "acrylic": "premium acrylic with UV-printed design",
    "pvc": "durable PVC with UV-printed design",
}

# Mounting type descriptions for content generation
MOUNTING_TYPES = {
    "self_adhesive": {
        "description": "Self-adhesive backing (peel and stick) - NO drilling required",
        "title_suffix": "Self-Adhesive",
        "bullet_point": "Self-adhesive backing for easy no-drill installation",
    },
    "pre_drilled": {
        "description": "Pre-drilled 4mm fixing holes for easy screw mounting",
        "title_suffix": "Pre-Drilled",
        "bullet_point": "Pre-drilled holes for easy screw mounting",
    },
    "magnetic": {
        "description": "Magnetic backing for repositionable mounting on metal surfaces",
        "title_suffix": "Magnetic",
        "bullet_point": "Magnetic backing for repositionable mounting",
    },
    "suction": {
        "description": "Suction cup mounting for glass and smooth surfaces",
        "title_suffix": "Suction Mount",
        "bullet_point": "Suction cup for easy glass/window mounting",
    },
}


@dataclass
class ProductData:
    """Product data for Amazon listing generation."""
    m_number: str
    description: str
    size: str
    color: str
    text_lines: list[str]
    material: str = "1mm_aluminium"
    mounting_type: str = "self_adhesive"
    image_urls: list[str] = field(default_factory=list)
    ean: str = ""  # EAN barcode from GS1
    
    @property
    def mounting_info(self) -> dict:
        """Get mounting type info."""
        return MOUNTING_TYPES.get(self.mounting_type.lower(), MOUNTING_TYPES["self_adhesive"])
    
    @property
    def size_cm(self) -> tuple[float, float]:
        """Get size dimensions in cm."""
        return SIZE_DIMENSIONS_CM.get(self.size.lower(), (11.5, 9.5))
    
    @property
    def material_display(self) -> str:
        """Get display description for material."""
        return MATERIALS.get(self.material.lower(), self.material)
    
    @property
    def color_display(self) -> str:
        """Get display name for color."""
        return COLOR_DISPLAY_NAMES.get(self.color.lower(), self.color.title())
    
    @property
    def size_display(self) -> str:
        """Get display name for size (without dimensions, for flatfile)."""
        return SIZE_DISPLAY_NAMES.get(self.size.lower(), self.size.title())
    
    @property
    def size_display_full(self) -> str:
        """Get display name for size with dimensions (for content generation)."""
        return SIZE_DISPLAY_NAMES_FULL.get(self.size.lower(), self.size.title())
    
    @property
    def size_map(self) -> str:
        """Get Amazon size_map value (XS, S, M, L, XL, etc.)."""
        return SIZE_MAP_VALUES.get(self.size.lower(), "M")


def generate_content_with_claude(
    product: ProductData,
    api_key: str,
    brand_name: str = "NorthByNorthEast",
    theme: str = "",
    use_cases: str = "",
) -> AmazonContent:
    """
    Generate Amazon SEO-optimized content using Claude API.
    
    Args:
        product: Product data
        api_key: Anthropic API key
        brand_name: Brand name for listings
        theme: Human-provided signage theme/description
        use_cases: Target use cases for the signage
        
    Returns:
        AmazonContent with title, description, bullets, and search terms
    """
    client = anthropic.Anthropic(api_key=api_key)
    
    # Use human-provided theme if available, otherwise extract from product data
    if theme:
        sign_text = theme
    else:
        sign_text = " ".join([t for t in product.text_lines if t])
        # If no text lines, extract sign type from description
        if not sign_text and product.description:
            desc = product.description
            for suffix in [" Sign Self Adhesive", " Sign Screw Mount", " Sign", " Self Adhesive", " Screw Mount"]:
                if desc.endswith(suffix):
                    desc = desc[:-len(suffix)]
                    break
            sign_text = desc
    
    length_cm, width_cm = product.size_cm
    
    mounting = product.mounting_info
    
    # Build use cases string
    use_cases_str = use_cases if use_cases else "offices, warehouses, car parks, shops, public spaces"
    
    prompt = f"""Generate Amazon UK product listing content for a sign product.

PRODUCT DETAILS:
- Sign Theme/Message: "{sign_text}"
- Size: {length_cm} x {width_cm} cm
- Color/Finish: {product.color_display}
- Material: {product.material_display}
- Mounting: {mounting["description"]}
- Features: Weatherproof, UV-resistant print, rounded corners
- Target Use Cases: {use_cases_str}

REQUIREMENTS:
1. TITLE (max 200 characters): Include primary keyword, key features, size in cm. Format: "[Sign Text] Sign – [dimensions]cm [Material], Weatherproof, {mounting["title_suffix"]}". Do NOT include brand name in title. Use dimensions (e.g. "15x10cm") NOT product code names.

2. DESCRIPTION (150-300 words): Detailed, persuasive product description. Emphasise the mounting method ({mounting["description"]}). Include material, dimensions, features, and use cases.

3. BULLET POINTS (exactly 5): Benefit-focused, keyword-rich. Each bullet should be 150-250 characters. Cover:
   - Material quality and finish
   - UV printing durability
   - {mounting["bullet_point"]}
   - Weatherproof construction
   - Clear messaging/visibility

4. SEARCH TERMS (max 250 characters): Backend keywords separated by spaces. Do NOT repeat words from title. Include synonyms, related terms, misspellings.

Respond in JSON format:
{{
    "title": "...",
    "description": "...",
    "bullet_points": ["...", "...", "...", "...", "..."],
    "search_terms": "..."
}}"""

    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=1500,
        messages=[
            {"role": "user", "content": prompt}
        ]
    )
    
    # Parse the response
    response_text = message.content[0].text
    
    # Extract JSON from response (handle markdown code blocks)
    json_match = re.search(r'\{[\s\S]*\}', response_text)
    if not json_match:
        raise ValueError(f"Could not parse JSON from Claude response: {response_text[:200]}")
    
    data = json.loads(json_match.group())
    
    return AmazonContent(
        title=data["title"][:200],  # Enforce max length
        description=data["description"],
        bullet_points=data["bullet_points"][:5],  # Max 5 bullets
        search_terms=data["search_terms"][:250],  # Enforce max length
    )


def convert_png_to_jpeg_local(png_path: Path, background_color=(255, 255, 255)) -> Path:
    """
    Convert PNG with transparency to JPEG with solid background.
    
    Args:
        png_path: Path to input PNG file
        background_color: RGB tuple for background (default white)
    
    Returns:
        Path to output JPEG file (same directory, .jpg extension)
    """
    from PIL import Image
    
    output_path = png_path.with_suffix(".jpg")
    
    img = Image.open(png_path)
    
    # Strip ICC profile and EXIF to avoid compatibility issues
    img.info.pop('icc_profile', None)
    
    if img.mode in ("RGBA", "LA", "P"):
        background = Image.new("RGB", img.size, background_color)
        if img.mode == "P":
            img = img.convert("RGBA")
        if img.mode == "RGBA":
            background.paste(img, mask=img.split()[3])
        else:
            background.paste(img)
        img = background
    elif img.mode != "RGB":
        img = img.convert("RGB")
    
    # Save with optimized baseline JPEG for Etsy compatibility
    img.save(output_path, "JPEG", quality=85, optimize=True)
    return output_path


def upload_to_cloudflare_r2(
    image_path: Path,
    bucket_name: str,
    account_id: str,
    access_key_id: str,
    secret_access_key: str,
    public_url_base: str,
    also_upload_jpeg: bool = False,
) -> str:
    """
    Upload an image to Cloudflare R2 and return the public URL.
    
    Args:
        image_path: Path to the image file
        bucket_name: R2 bucket name
        account_id: Cloudflare account ID
        access_key_id: R2 access key ID
        secret_access_key: R2 secret access key
        public_url_base: Base URL for public access (e.g., https://images.yoursite.com)
        also_upload_jpeg: If True and image is PNG, also convert and upload JPEG version
        
    Returns:
        Public URL of the uploaded image
    """
    # Configure S3 client for R2 (must use 'auto' region)
    s3_client = boto3.client(
        "s3",
        endpoint_url=f"https://{account_id}.r2.cloudflarestorage.com",
        aws_access_key_id=access_key_id,
        aws_secret_access_key=secret_access_key,
        region_name="auto",
        config=Config(signature_version="s3v4"),
    )
    
    # Upload the original file
    object_key = image_path.name
    content_type = "image/png" if image_path.suffix.lower() == ".png" else "image/jpeg"
    
    s3_client.upload_file(
        str(image_path),
        bucket_name,
        object_key,
        ExtraArgs={"ContentType": content_type},
    )
    
    public_url = f"{public_url_base.rstrip('/')}/{object_key}"
    logging.info("Uploaded %s to %s", image_path.name, public_url)
    
    # Also upload JPEG version if requested and source is PNG
    if also_upload_jpeg and image_path.suffix.lower() == ".png":
        try:
            jpeg_path = convert_png_to_jpeg_local(image_path)
            jpeg_key = jpeg_path.name
            s3_client.upload_file(
                str(jpeg_path),
                bucket_name,
                jpeg_key,
                ExtraArgs={"ContentType": "image/jpeg"},
            )
            logging.info("Also uploaded JPEG: %s", jpeg_key)
            # Clean up temp JPEG
            jpeg_path.unlink(missing_ok=True)
        except Exception as e:
            logging.warning("Failed to create JPEG version: %s", e)
    
    return public_url


def upload_single_image(args_tuple) -> tuple[int, str]:
    """
    Upload a single image - wrapper for parallel execution.
    Returns (sort_index, url) to maintain order.
    """
    idx, img_file, bucket_name, account_id, access_key_id, secret_access_key, public_url_base, also_upload_jpeg = args_tuple
    url = upload_to_cloudflare_r2(
        img_file,
        bucket_name,
        account_id,
        access_key_id,
        secret_access_key,
        public_url_base,
        also_upload_jpeg=also_upload_jpeg,
    )
    return (idx, url)


def upload_product_images(
    m_folder: Path,
    bucket_name: str,
    account_id: str,
    access_key_id: str,
    secret_access_key: str,
    public_url_base: str,
    also_upload_jpeg: bool = True,
) -> list[str]:
    """
    Upload all images from a product's M Number folder to R2 using parallel processing.
    
    Args:
        also_upload_jpeg: If True, also upload JPEG versions for Etsy compatibility
    
    Returns list of URLs in order: main (001), dimensions (002), peel_and_stick (003), rear (004), etc.
    """
    images_dir = m_folder / "002 Images"
    if not images_dir.exists():
        logging.warning("Images directory not found: %s", images_dir)
        return []
    
    # Get sorted list of images
    img_files = sorted(images_dir.glob("*.png"))
    if not img_files:
        return []
    
    # Prepare arguments for parallel upload
    upload_args = [
        (idx, img_file, bucket_name, account_id, access_key_id, secret_access_key, public_url_base, also_upload_jpeg)
        for idx, img_file in enumerate(img_files)
    ]
    
    # Upload in parallel
    results = [None] * len(img_files)
    with ThreadPoolExecutor(max_workers=min(MAX_UPLOAD_WORKERS, len(img_files))) as executor:
        futures = {executor.submit(upload_single_image, args): args[0] for args in upload_args}
        for future in as_completed(futures):
            idx, url = future.result()
            results[idx] = url
    
    return results


def read_products_from_csv(csv_path: Path, qa_filter: str = "approved") -> list[ProductData]:
    """
    Read products from CSV file.
    
    Args:
        csv_path: Path to CSV file
        qa_filter: Only include products with this qa_status. 
                   Use "all" to include all products regardless of status.
    """
    products = []
    skipped = 0
    
    with csv_path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            m_number = (row.get("m_number") or "").strip()
            if not m_number:
                continue
            
            # Check QA status filter
            qa_status = (row.get("qa_status") or "pending").strip().lower()
            if qa_filter != "all" and qa_status != qa_filter:
                skipped += 1
                continue
                
            text_lines = [
                (row.get("text_line_1") or "").strip(),
                (row.get("text_line_2") or "").strip(),
                (row.get("text_line_3") or "").strip(),
            ]
            
            material = (row.get("material") or "1mm_aluminium").strip().lower()
            mounting_type = (row.get("mounting_type") or "self_adhesive").strip().lower()
            
            products.append(ProductData(
                m_number=m_number,
                description=(row.get("description") or "").strip(),
                size=(row.get("size") or "saville").strip().lower(),
                color=(row.get("color") or "silver").strip().lower(),
                text_lines=text_lines,
                material=material,
                mounting_type=mounting_type,
                ean=(row.get("ean") or "").strip(),
            ))
    
    if skipped > 0:
        logging.info("Skipped %d products (qa_status != '%s')", skipped, qa_filter)
    
    return products


def derive_parent_sku_from_description(description: str) -> str:
    """
    Derive parent SKU name from product description.
    E.g., "Self Adhesive no public access aluminium sign" -> "NO_PUBLIC_ACCESS_PARENT"
    """
    # Extract the sign type from description
    desc_lower = description.lower()
    
    # Remove common prefixes
    for prefix in ["self adhesive ", "pre-drilled ", "magnetic ", "suction "]:
        if desc_lower.startswith(prefix):
            desc_lower = desc_lower[len(prefix):]
    
    # Remove common suffixes
    for suffix in [" aluminium sign", " aluminum sign", " sign", " metal sign"]:
        if desc_lower.endswith(suffix):
            desc_lower = desc_lower[:-len(suffix)]
    
    # Convert to uppercase with underscores
    parent_name = desc_lower.strip().upper().replace(" ", "_").replace("-", "_")
    
    # Clean up multiple underscores
    while "__" in parent_name:
        parent_name = parent_name.replace("__", "_")
    
    return f"{parent_name}_PARENT"


def generate_flatfile(
    products: list[ProductData],
    contents: dict[str, AmazonContent],
    output_path: Path,
    brand_name: str = "NorthByNorthEast",
    parent_sku: Optional[str] = None,
) -> None:
    """
    Generate Amazon flatfile XLSX in exact Amazon upload format.
    
    The format matches Amazon's Signage template with:
    - Row 1: Template metadata
    - Row 2: Human-readable column labels
    - Row 3: Attribute names (what Amazon uses internally)
    - Row 4+: Data (Parent row first, then Child rows)
    
    Args:
        products: List of product data
        contents: Dict mapping m_number to AmazonContent
        output_path: Output XLSX file path
        brand_name: Brand name for listings
        parent_sku: Optional parent SKU for variations (auto-derived if not provided)
    """
    # Amazon template metadata for Row 1
    TEMPLATE_METADATA = [
        "TemplateType=fptcustom",
        "Version=2025.1207",
        "TemplateSignature=U0lHTkFHRQ==",  # Base64 for "SIGNAGE"
        # Settings string contains marketplace IDs, browse classifications, etc.
        "settings=attributeRow=3&dataRow=4&feedType=610841&contentLanguageTag=en_GB&headerLanguageTag=en_GB&labelRow=2&primaryMarketplaceId=amzn1.mp.o.A1F83G8C2ARO7P&productTypeRequirement=LISTING&ptds=U0lHTkFHRQ%3D%3D",
    ]
    
    # Define all columns in exact Amazon order
    # Format: (attribute_name, label, group_header)
    AMAZON_COLUMNS = [
        # Core product info
        ("feed_product_type", "Product Type", None),
        ("item_sku", "Seller SKU", None),
        ("update_delete", "Update Delete", None),
        ("brand_name", "Brand Name", None),
        ("external_product_id", "Product ID", None),
        ("external_product_id_type", "Product ID Type", None),
        ("product_description", "Product Description", None),
        ("part_number", "Manufacturer Part Number", None),
        ("manufacturer", "Manufacturer", None),
        ("item_name", "Item Name (aka Title)", None),
        ("language_value", "Language", None),
        ("recommended_browse_nodes", "Recommended Browse Nodes", None),
        # Images
        ("main_image_url", "Main Image URL", None),
        ("other_image_url1", "Other Image Url1", None),
        ("other_image_url2", "Other Image Url2", None),
        ("other_image_url3", "Other Image Url3", None),
        ("other_image_url4", "Other Image Url4", None),
        ("other_image_url5", "Other Image Url5", None),
        ("other_image_url6", "Other Image Url6", None),
        ("other_image_url7", "Other Image Url7", None),
        ("other_image_url8", "Other Image Url8", None),
        # Variation
        ("relationship_type", "Relationship Type", "Variation"),
        ("variation_theme", "Variation Theme", None),
        ("parent_sku", "Parent SKU", None),
        ("parent_child", "Parentage", None),
        # Discovery
        ("style_name", "Style Name", "Discovery"),
        ("bullet_point1", "Key Product Features", None),
        ("bullet_point2", "Key Product Features", None),
        ("bullet_point3", "Key Product Features", None),
        ("bullet_point4", "Key Product Features", None),
        ("bullet_point5", "Key Product Features", None),
        ("generic_keywords", "Search Terms", None),
        ("color_name", "Colour", None),
        ("size_name", "Size", None),
        ("color_map", "Colour Map", None),
        # Dimensions
        ("size_map", "Size Map", "Dimensions"),
        ("length_longer_edge", "Item Length Longer Edge", None),
        ("length_longer_edge_unit_of_measure", "Item Length Unit", None),
        ("width_shorter_edge", "Item Width Shorter Edge", None),
        ("width_shorter_edge_unit_of_measure", "Item Width Unit", None),
        # Compliance
        ("batteries_required", "Is this product a battery or does it utilise batteries?", "Fulfillment"),
        ("country_of_origin", "Country/Region Of Origin", None),
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template"
    
    # Row 1: Template metadata
    for col, meta in enumerate(TEMPLATE_METADATA, 1):
        ws.cell(row=1, column=col, value=meta)
    
    # Row 2: Human-readable labels
    for col, (_, label, _) in enumerate(AMAZON_COLUMNS, 1):
        ws.cell(row=2, column=col, value=label)
    
    # Row 3: Attribute names
    for col, (attr, _, _) in enumerate(AMAZON_COLUMNS, 1):
        ws.cell(row=3, column=col, value=attr)
    
    # Auto-derive parent SKU from first product's description if not provided
    if not parent_sku and products:
        parent_sku = derive_parent_sku_from_description(products[0].description)
        logging.info("Auto-derived parent SKU: %s", parent_sku)
    
    # Create a parent title (without size)
    parent_title = ""
    if products and products[0].m_number in contents:
        # Extract base title without size dimensions
        child_title = contents[products[0].m_number].title
        # Remove size like "9.5x9.5cm" or "11x9.5cm" from title
        import re
        parent_title = re.sub(r'\s*–?\s*\d+\.?\d*x\d+\.?\d*\s*cm\s*', ' – ', child_title)
        parent_title = parent_title.replace('  ', ' ').strip()
    
    # Row 4: Parent row
    row_num = 4
    parent_data = {
        "feed_product_type": "signage",
        "item_sku": parent_sku,
        "update_delete": "Update",
        "brand_name": brand_name,
        "external_product_id": "",  # Parent has no EAN
        "external_product_id_type": "",
        "part_number": parent_sku,
        "item_name": parent_title,
        "recommended_browse_nodes": "330215031",  # Signage browse node
        "variation_theme": "Size & Colour",
        "parent_child": "Parent",
        "batteries_required": "No",
        "country_of_origin": "Great Britain",
    }
    
    for col, (attr, _, _) in enumerate(AMAZON_COLUMNS, 1):
        value = parent_data.get(attr, "")
        ws.cell(row=row_num, column=col, value=value)
    row_num += 1
    
    # Row 5+: Child products
    for product in products:
        content = contents.get(product.m_number)
        if not content:
            logging.warning("No content for %s, skipping", product.m_number)
            continue
        
        length_cm, width_cm = product.size_cm
        size_code = product.size_map  # XS, S, M, L, XL
        
        # Style name format: Color_SizeCode (e.g., "Silver_XS")
        style_name = f"{product.color_display}_{size_code}"
        
        # Build row data
        row_data = {
            "feed_product_type": "signage",
            "item_sku": product.m_number,
            "update_delete": "Update",
            "brand_name": brand_name,
            "external_product_id": product.ean,
            "external_product_id_type": "EAN" if product.ean else "",
            "product_description": content.description,
            "part_number": product.m_number,
            "manufacturer": brand_name,
            "item_name": content.title,
            "language_value": "en_GB",
            "recommended_browse_nodes": "330215031",
            "relationship_type": "Variation",
            "variation_theme": "Size & Colour",
            "parent_sku": parent_sku,
            "parent_child": "Child",
            "style_name": style_name,
            "bullet_point1": content.bullet_points[0] if len(content.bullet_points) > 0 else "",
            "bullet_point2": content.bullet_points[1] if len(content.bullet_points) > 1 else "",
            "bullet_point3": content.bullet_points[2] if len(content.bullet_points) > 2 else "",
            "bullet_point4": content.bullet_points[3] if len(content.bullet_points) > 3 else "",
            "bullet_point5": content.bullet_points[4] if len(content.bullet_points) > 4 else "",
            "generic_keywords": content.search_terms,
            "color_name": product.color_display,
            "size_name": size_code,  # Use XS, S, M, L, XL
            "color_map": product.color_display,
            "size_map": size_code,
            "length_longer_edge": str(length_cm),
            "length_longer_edge_unit_of_measure": "Centimetres",
            "width_shorter_edge": str(width_cm),
            "width_shorter_edge_unit_of_measure": "Centimetres",
            "batteries_required": "No",
            "country_of_origin": "Great Britain",
        }
        
        # Add image URLs if available
        if product.image_urls:
            row_data["main_image_url"] = product.image_urls[0] if len(product.image_urls) > 0 else ""
            row_data["other_image_url1"] = product.image_urls[1] if len(product.image_urls) > 1 else ""
            row_data["other_image_url2"] = product.image_urls[2] if len(product.image_urls) > 2 else ""
            row_data["other_image_url3"] = product.image_urls[3] if len(product.image_urls) > 3 else ""
            row_data["other_image_url4"] = product.image_urls[4] if len(product.image_urls) > 4 else ""
            row_data["other_image_url5"] = product.image_urls[5] if len(product.image_urls) > 5 else ""
            row_data["other_image_url6"] = product.image_urls[6] if len(product.image_urls) > 6 else ""
            row_data["other_image_url7"] = product.image_urls[7] if len(product.image_urls) > 7 else ""
            row_data["other_image_url8"] = product.image_urls[8] if len(product.image_urls) > 8 else ""
        
        # Write to row
        for col, (attr, _, _) in enumerate(AMAZON_COLUMNS, 1):
            value = row_data.get(attr, "")
            ws.cell(row=row_num, column=col, value=value)
        
        row_num += 1
    
    # Auto-adjust column widths
    for col in range(1, len(AMAZON_COLUMNS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    wb.save(output_path)
    logging.info("Saved Amazon flatfile to %s with %d products (parent + %d children)", 
                 output_path, row_num - 4, row_num - 5)


def main():
    parser = argparse.ArgumentParser(description="Generate Amazon UK flatfile content")
    parser.add_argument("--csv", type=Path, default=Path("products_test.csv"), help="Input CSV file")
    parser.add_argument("--exports", type=Path, default=Path("exports"), help="Exports directory with M folders")
    parser.add_argument("--output", type=Path, default=Path("amazon_flatfile.xlsx"), help="Output flatfile path")
    parser.add_argument("--brand", type=str, default="NorthByNorthEast", help="Brand name")
    parser.add_argument("--parent-sku", type=str, default=None, help="Parent SKU for variations")
    parser.add_argument("--upload-images", action="store_true", help="Upload images to Cloudflare R2")
    parser.add_argument("--dry-run", action="store_true", help="Generate content without saving")
    parser.add_argument("--qa-filter", type=str, default="approved", help="QA status filter (approved/pending/rejected/all)")
    parser.add_argument("--theme", type=str, default="", help="Human-provided signage theme/description for AI context")
    parser.add_argument("--use-cases", type=str, default="", help="Target use cases (e.g., parks, offices, warehouses)")
    parser.add_argument("--theme-file", type=Path, default=None, help="File containing theme text (alternative to --theme)")
    parser.add_argument("--use-cases-file", type=Path, default=None, help="File containing use cases text (alternative to --use-cases)")
    parser.add_argument("--m-number", type=str, default=None, help="Process only a specific M number (e.g., M1220)")
    args = parser.parse_args()
    
    # Read theme/use-cases from files if specified
    if args.theme_file and args.theme_file.exists():
        args.theme = args.theme_file.read_text(encoding="utf-8").strip()
        logging.info("Loaded theme from file: %s", args.theme[:50] + "..." if len(args.theme) > 50 else args.theme)
    if args.use_cases_file and args.use_cases_file.exists():
        args.use_cases = args.use_cases_file.read_text(encoding="utf-8").strip()
        logging.info("Loaded use cases from file: %s", args.use_cases[:50] + "..." if len(args.use_cases) > 50 else args.use_cases)
    
    # Get API key from environment
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        logging.error("ANTHROPIC_API_KEY environment variable not set")
        return 1
    
    # Read products (filtered by QA status)
    products = read_products_from_csv(args.csv, qa_filter=args.qa_filter)
    logging.info("Loaded %d products from %s (qa_filter=%s)", len(products), args.csv, args.qa_filter)
    
    # Filter to specific M number if requested
    if args.m_number:
        products = [p for p in products if p.m_number == args.m_number]
        if not products:
            logging.error("M number %s not found in CSV", args.m_number)
            return 1
        logging.info("Filtered to M number: %s", args.m_number)
    
    # Generate content for each product
    contents = {}
    for product in products:
        logging.info("Generating content for %s...", product.m_number)
        try:
            content = generate_content_with_claude(product, api_key, args.brand, args.theme, args.use_cases)
            contents[product.m_number] = content
            logging.info("  Title: %s", content.title[:80] + "..." if len(content.title) > 80 else content.title)
        except Exception as e:
            logging.error("Failed to generate content for %s: %s", product.m_number, e)
    
    # Upload images if requested
    if args.upload_images:
        r2_account_id = os.environ.get("R2_ACCOUNT_ID")
        r2_access_key = os.environ.get("R2_ACCESS_KEY_ID")
        r2_secret_key = os.environ.get("R2_SECRET_ACCESS_KEY")
        r2_bucket = os.environ.get("R2_BUCKET_NAME")
        r2_public_url = os.environ.get("R2_PUBLIC_URL")
        
        if not all([r2_account_id, r2_access_key, r2_secret_key, r2_bucket, r2_public_url]):
            logging.error("R2 environment variables not fully set")
            return 1
        
        # Collect all images from all products for parallel upload
        all_upload_tasks = []
        for product in products:
            m_folder_name = f"{product.m_number} {product.description} {product.color_display} {product.size.title()}"
            m_folder = args.exports / m_folder_name
            
            if m_folder.exists():
                images_dir = m_folder / "002 Images"
                if images_dir.exists():
                    img_files = sorted(images_dir.glob("*.png"))
                    for idx, img_file in enumerate(img_files):
                        all_upload_tasks.append((product.m_number, idx, img_file))
            else:
                logging.warning("M folder not found: %s", m_folder)
        
        # Upload all images in parallel
        logging.info("Uploading %d images using %d parallel workers...", len(all_upload_tasks), MAX_UPLOAD_WORKERS)
        upload_results = {}  # {m_number: {idx: url}}
        completed = 0
        
        def upload_task(task):
            m_number, idx, img_file = task
            url = upload_to_cloudflare_r2(
                img_file, r2_bucket, r2_account_id,
                r2_access_key, r2_secret_key, r2_public_url,
                also_upload_jpeg=True
            )
            return (m_number, idx, url)
        
        with ThreadPoolExecutor(max_workers=MAX_UPLOAD_WORKERS) as executor:
            futures = {executor.submit(upload_task, task): task for task in all_upload_tasks}
            for future in as_completed(futures):
                m_number, idx, url = future.result()
                if m_number not in upload_results:
                    upload_results[m_number] = {}
                upload_results[m_number][idx] = url
                completed += 1
                if completed % 10 == 0 or completed == len(all_upload_tasks):
                    logging.info("Upload progress: %d/%d images", completed, len(all_upload_tasks))
        
        # Assign URLs back to products in correct order
        for product in products:
            if product.m_number in upload_results:
                results = upload_results[product.m_number]
                product.image_urls = [results[i] for i in sorted(results.keys())]
    
    # Generate flatfile
    if not args.dry_run:
        generate_flatfile(products, contents, args.output, args.brand, args.parent_sku)
    
    logging.info("Done! Generated content for %d products", len(contents))
    return 0


if __name__ == "__main__":
    exit(main())
