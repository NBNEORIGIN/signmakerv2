#!/usr/bin/env python3
"""
Etsy CSV Lister Generator

Reads Amazon flatfile XLSM and generates Etsy CSV Lister compatible XLSX.
Supports 2-variation listings (Size & Color).
"""

import argparse
import logging
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
)

# CSV Lister template configuration (from your Config sheet)
ETSY_CONFIG = {
    "shop_id": 11706740,
    "shop_name": "NorthByNorthEastSign",
    "shipping_profile_id": 208230423243,  # Postage 2025
    "processing_profile_id": 1402336022581,  # 1-2 days
    "return_policy_id": 1074420280634,  # 30 days
    "shop_section_id": 23027026,  # Business Signage
    "category_id": 2844,  # Signs
}

# Size mappings from Amazon size codes to Etsy display
SIZE_DISPLAY = {
    "S": "11.5 x 9.5 cm",
    "M": "15 x 10 cm",
    "L": "21 x 14.8 cm",
    "XL": "11.5 x 9.5 cm (Portrait)",
    "XS": "11.5 x 9.5 cm",
    "XXS": "11.5 x 9.5 cm",
}

# Size dimensions for attributes
SIZE_DIMENSIONS = {
    "S": (11.5, 9.5),
    "M": (15.0, 10.0),
    "L": (21.0, 14.8),
    "XL": (11.5, 9.5),
    "XS": (11.5, 9.5),
}

# CSV Lister column headers (Row 2 in template)
CSV_LISTER_COLUMNS = [
    "title",
    "description",
    "variation_1_type",
    "variation_1_value",
    "variation_2_type",
    "variation_2_value",
    "price",
    "quantity",
    "readiness_state_id",
    "sku",
    "who_made",
    "when_made",
    "is_supply",
    "should_auto_renew",
    "shipping_profile_id",
    "materials",
    "tags",
    "image_url_1",
    "image_url_2",
    "image_url_3",
    "image_url_4",
    "image_url_5",
    "image_url_6",
    "image_url_7",
    "image_url_8",
    "image_url_9",
    "image_url_10",
    "image_1_alt_text",
    "image_2_alt_text",
    "image_3_alt_text",
    "image_4_alt_text",
    "image_5_alt_text",
    "image_6_alt_text",
    "image_7_alt_text",
    "image_8_alt_text",
    "image_9_alt_text",
    "image_10_alt_text",
    "video_url",
    "shop_section_id",
    "return_policy_id",
    "processing_min",
    "processing_max",
    "is_personalizable",
    "personalization_is_required",
    "personalization_char_count_max",
    "personalization_instructions",
    "is_customizable",
    "is_taxable",
    "item_weight",
    "item_weight_unit",
    "item_length",
    "item_width",
    "item_height",
    "item_dimensions_unit",
    "attribute_materials_1",
    "attribute_materials_2",
    "attribute_materials_3",
    "attribute_materials_4",
    "attribute_materials_5",
    "attribute_primary_color",
    "attribute_secondary_color",
    "attribute_width",
    "attribute_width_scale",
    "attribute_height",
    "attribute_height_scale",
    "attribute_depth",
    "attribute_depth_scale",
    "attribute_illuminated",
    "attribute_sustainability_1",
    "attribute_sustainability_2",
    "attribute_shape",
    "attribute_orientation",
    "attribute_framing",
    "attribute_bulb_type",
    "attribute_mount_type",
    "attribute_style",
    "attribute_occasion",
    "attribute_holiday",
    "attribute_room_1",
    "attribute_room_2",
    "attribute_room_3",
    "attribute_room_4",
    "attribute_room_5",
    "attribute_season",
    "attribute_graphic",
    "attribute_can_be_personalized",
    "attribute_diameter",
    "attribute_diameter_scale",
    "attribute_dimensions",
    "attribute_dimensions_scale",
    "attribute_fabric",
    "attribute_finish",
    "attribute_flavor",
    "attribute_length",
    "attribute_length_scale",
    "attribute_material",
    "attribute_pattern",
    "attribute_scent",
    "attribute_size",
    "attribute_size_scale",
    "attribute_style",
    "attribute_weight",
    "attribute_weight_scale",
    "attribute_device",
]


def read_amazon_flatfile(flatfile_path: Path) -> list[dict]:
    """
    Read Amazon flatfile XLSM and extract child product data.
    
    Returns list of product dicts with relevant fields.
    """
    wb = openpyxl.load_workbook(flatfile_path, read_only=True, data_only=True)
    ws = wb["Template"]
    
    # Build column index from row 3 (attribute names)
    col_index = {}
    for col in range(1, ws.max_column + 1):
        attr = ws.cell(3, col).value
        if attr:
            col_index[attr] = col
    
    products = []
    
    # Read data rows (starting from row 4)
    for row in range(4, ws.max_row + 1):
        sku = ws.cell(row, col_index.get("item_sku", 2)).value
        parent_child = ws.cell(row, col_index.get("parent_child", 32)).value
        
        if not sku or parent_child != "Child":
            continue
        
        product = {
            "sku": sku,
            "title": ws.cell(row, col_index.get("item_name", 10)).value,
            "description": ws.cell(row, col_index.get("product_description", 7)).value,
            "color": ws.cell(row, col_index.get("color_name", 45)).value,
            "size": ws.cell(row, col_index.get("size_name", 46)).value,
            "main_image": ws.cell(row, col_index.get("main_image_url", 13)).value,
            "image_2": ws.cell(row, col_index.get("other_image_url1", 14)).value,
            "image_3": ws.cell(row, col_index.get("other_image_url2", 15)).value,
            "image_4": ws.cell(row, col_index.get("other_image_url3", 16)).value,
            "image_5": ws.cell(row, col_index.get("other_image_url4", 17)).value,
            "bullets": [],
            "keywords": ws.cell(row, col_index.get("generic_keywords", 44)).value,
        }
        
        # Collect bullet points
        for i in range(1, 6):
            bp_col = col_index.get(f"bullet_point{i}")
            if bp_col:
                bp = ws.cell(row, bp_col).value
                if bp:
                    product["bullets"].append(bp)
        
        products.append(product)
    
    wb.close()
    logging.info("Read %d child products from %s", len(products), flatfile_path.name)
    return products


def convert_amazon_to_etsy_title(amazon_title: str) -> str:
    """
    Convert Amazon title to Etsy format (max 140 chars).
    """
    if not amazon_title:
        return ""
    
    # Etsy max is 140 chars
    title = amazon_title[:140]
    return title


def convert_description_to_etsy(description: str, bullets: list[str]) -> str:
    """
    Convert Amazon description + bullets to Etsy description format.
    """
    parts = []
    
    if description:
        parts.append(description)
    
    if bullets:
        parts.append("\n\n✦ KEY FEATURES ✦\n")
        for bullet in bullets:
            parts.append(f"• {bullet}")
    
    parts.append("\n\n✦ SHIPPING ✦")
    parts.append("Made to order. Dispatched within 1-2 business days.")
    
    return "\n".join(parts)


def convert_keywords_to_tags(keywords: str) -> str:
    """
    Convert Amazon keywords to Etsy tags (max 13 tags, each max 20 chars).
    """
    if not keywords:
        return ""
    
    # Split by spaces and common separators
    words = keywords.replace(",", " ").split()
    
    # Build tags, combining short words
    tags = []
    current_tag = ""
    
    for word in words:
        word = word.strip()
        if not word:
            continue
        
        if len(current_tag) == 0:
            current_tag = word
        elif len(current_tag) + 1 + len(word) <= 20:
            current_tag += " " + word
        else:
            if current_tag:
                tags.append(current_tag[:20])
            current_tag = word
        
        if len(tags) >= 13:
            break
    
    if current_tag and len(tags) < 13:
        tags.append(current_tag[:20])
    
    return ",".join(tags[:13])


def generate_etsy_csv_lister(
    products: list[dict],
    output_path: Path,
    template_path: Path,
    price: float = 9.99,
    quantity: int = 999,
) -> None:
    """
    Generate Etsy CSV Lister XLSX from product data.
    Creates a fresh workbook to avoid Excel corruption issues.
    """
    # Create fresh workbook
    wb = openpyxl.Workbook()
    
    # Load template to get Config data
    template_wb = openpyxl.load_workbook(template_path, data_only=True)
    template_config = template_wb["Config"]
    
    # Create Config sheet (copy from template)
    config = wb.active
    config.title = "Config"
    for row in template_config.iter_rows(min_row=1, max_row=template_config.max_row):
        for cell in row:
            config.cell(cell.row, cell.column, cell.value)
    
    # Create Cat_1_Signs sheet
    ws = wb.create_sheet("Cat_1_Signs")
    
    # Row 1: Category metadata (CRITICAL for CSV Lister validation)
    ws["A1"] = "Category: Home & Living > Home Decor > Wall Decor > Wall Hangings > Signs"
    ws["D1"] = "Item Type: physical"
    ws["F1"] = "Variations: 2"
    ws["H1"] = "Category ID:"
    ws["I1"] = 2844  # Category ID as separate numeric cell
    ws["J1"] = "Template Type: standard"
    
    # Row 2: Column headers
    for col_idx, col_name in enumerate(CSV_LISTER_COLUMNS, 1):
        ws.cell(row=2, column=col_idx, value=col_name)
    
    # Row 3: Skip (descriptions row - leave empty for cleaner file)
    
    # Data starts at row 4
    row_num = 4
    
    template_wb.close()
    
    for product in products:
        size_display = SIZE_DISPLAY.get(product["size"], product["size"])
        dimensions = SIZE_DIMENSIONS.get(product["size"], (11.5, 9.5))
        
        # Build row data
        row_data = {
            "title": convert_amazon_to_etsy_title(product["title"]),
            "description": convert_description_to_etsy(product["description"], product["bullets"]),
            "variation_1_type": "Size",
            "variation_1_value": size_display,
            "variation_2_type": "Primary color",
            "variation_2_value": product["color"],
            "price": price,
            "quantity": quantity,
            "readiness_state_id": int(ETSY_CONFIG["processing_profile_id"]),  # Must be integer
            "sku": product["sku"],
            "who_made": "i_did",
            "when_made": "2020_2025",
            "is_supply": "FALSE",
            "should_auto_renew": "TRUE",
            "shipping_profile_id": int(ETSY_CONFIG["shipping_profile_id"]),  # Must be integer
            "materials": "Brushed Aluminium,UV Print,Self-Adhesive Backing",
            "tags": convert_keywords_to_tags(product["keywords"]),
            "image_url_1": product["main_image"] or "",
            "image_url_2": product["image_2"] or "",
            "image_url_3": product["image_3"] or "",
            "image_url_4": product["image_4"] or "",
            "image_url_5": product["image_5"] or "",
            "image_1_alt_text": f"{product['title'][:100]}" if product["title"] else "",
            "shop_section_id": int(ETSY_CONFIG["shop_section_id"]),
            "return_policy_id": int(ETSY_CONFIG["return_policy_id"]),
            "processing_min": 1,
            "processing_max": 2,
            "is_personalizable": "FALSE",
            "is_customizable": "FALSE",
            "is_taxable": "TRUE",
            "item_length": dimensions[0],
            "item_width": dimensions[1],
            "item_height": 0.1,
            "item_dimensions_unit": "cm",
            "attribute_materials_1": "Aluminium",
            "attribute_materials_2": "UV Print",
            "attribute_primary_color": product["color"],
            "attribute_width": dimensions[1],
            "attribute_width_scale": "cm",
            "attribute_height": dimensions[0],
            "attribute_height_scale": "cm",
            "attribute_mount_type": "Self-adhesive",
            "attribute_style": "Modern",
            "attribute_room_1": "Office",
            "attribute_room_2": "Entryway",
        }
        
        # Write to row
        for col_idx, col_name in enumerate(CSV_LISTER_COLUMNS, 1):
            value = row_data.get(col_name, "")
            ws.cell(row=row_num, column=col_idx, value=value)
        
        row_num += 1
    
    # Save output
    wb.save(output_path)
    wb.close()
    
    logging.info("Generated %d listings to %s", len(products), output_path)


def main():
    parser = argparse.ArgumentParser(description="Generate Etsy CSV Lister from Amazon flatfile")
    parser.add_argument("--input", type=Path, required=True, help="Input Amazon flatfile XLSM")
    parser.add_argument("--output", type=Path, default=None, help="Output Etsy CSV Lister XLSX")
    parser.add_argument("--template", type=Path, 
                       default=Path(__file__).parent / "003 FLATFILES" / "etsy_csv_lister_template.xlsx",
                       help="CSV Lister template XLSX")
    parser.add_argument("--price", type=float, default=9.99, help="Listing price in GBP")
    parser.add_argument("--quantity", type=int, default=999, help="Stock quantity")
    args = parser.parse_args()
    
    if not args.input.exists():
        logging.error("Input file not found: %s", args.input)
        return 1
    
    if not args.template.exists():
        logging.error("Template file not found: %s", args.template)
        return 1
    
    # Default output name
    if args.output is None:
        args.output = args.input.parent / f"{args.input.stem}_etsy.xlsx"
    
    # Read Amazon flatfile
    products = read_amazon_flatfile(args.input)
    
    if not products:
        logging.error("No products found in flatfile")
        return 1
    
    # Generate Etsy CSV Lister
    generate_etsy_csv_lister(
        products,
        args.output,
        args.template,
        price=args.price,
        quantity=args.quantity,
    )
    
    logging.info("Done!")
    return 0


if __name__ == "__main__":
    exit(main())
